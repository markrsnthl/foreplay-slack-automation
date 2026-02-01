#!/usr/bin/env python3
"""
MVR Digital - 2026 Revenue Estimate Dashboard
Built from actual budget tracker data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============= SHEET 1: CLIENT CONTRACTS =============
contracts = wb.active
contracts.title = "2026 Contracts"

headers = ["Client", "Base Retainer", "% of Ad Spend", "Notes"]
for col, h in enumerate(headers, 1):
    cell = contracts.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")

# 2026 clients with contract terms (blue = editable inputs)
clients_config = [
    ["Peddle", 0, 0.055, "5.5% of ad spend"],
    ["SUNFLOW", 0, 0.07, "7% attributed revenue excl. branded"],
    ["Sonsie Skin", 4000, 0.08, "$4K base + 8% revenue share"],
    ["Kaspar & Lugay", 1000, 0.05, "$1K base + 5% of ad spend"],
    ["Wrensilva", 4000, 0.05, "$4K base + 5% attributed revenue"],
]

for row_idx, data in enumerate(clients_config, 2):
    for col_idx, val in enumerate(data, 1):
        cell = contracts.cell(row=row_idx, column=col_idx, value=val)
        if col_idx in [2, 3]:  # Editable inputs
            cell.font = Font(color="0000FF")
        if col_idx == 2:
            cell.number_format = '"$"#,##0'
        if col_idx == 3:
            cell.number_format = '0.0%'

contracts.column_dimensions['A'].width = 18
contracts.column_dimensions['B'].width = 14
contracts.column_dimensions['C'].width = 14
contracts.column_dimensions['D'].width = 35

# ============= SHEET 2: AD SPEND (from budget trackers) =============
spend = wb.create_sheet("2026 Ad Spend")

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
client_names = [c[0] for c in clients_config]

# Headers
spend.cell(row=1, column=1, value="Client").font = Font(bold=True)
for col, month in enumerate(months, 2):
    cell = spend.cell(row=1, column=col, value=month)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center")

cell = spend.cell(row=1, column=14, value="Total")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="1F4E79")

# Actual budget tracker data
budget_data = {
    "Peddle": [1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600, 1129600],  # Extrapolated Q1 for full year
    "SUNFLOW": [9000, 15000, 42000, 42000, 150000, 210000, 125000, 65000, 15000, 17000, 44000, 32000],
    "Sonsie Skin": [17667, 17656, 20000, 20000, 74000, 81000, 63000, 42600, 40000, 65000, 90000, 60000],
    "Kaspar & Lugay": [109880, 114056, 119067, 125081, 132297, 140956, 151347, 163817, 178780, 196736, 218283, 244140],
    "Wrensilva": [100000, 100000, 100000, 100000, 100000, 100000, 100000, 100000, 100000, 120000, 120000, 120000],
}

for row_idx, client in enumerate(client_names, 2):
    spend.cell(row=row_idx, column=1, value=client).font = Font(bold=True)

    data = budget_data.get(client, [0]*12)
    for col_idx, val in enumerate(data, 2):
        cell = spend.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(color="0000FF")
        cell.number_format = '"$"#,##0'

    # Total formula
    spend.cell(row=row_idx, column=14, value=f"=SUM(B{row_idx}:M{row_idx})")
    spend.cell(row=row_idx, column=14).number_format = '"$"#,##0'
    spend.cell(row=row_idx, column=14).font = Font(bold=True)

# Total row
total_row = len(client_names) + 2
spend.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
spend.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9E2F3")
for col in range(2, 15):
    spend.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
    spend.cell(row=total_row, column=col).number_format = '"$"#,##0'
    spend.cell(row=total_row, column=col).font = Font(bold=True)
    spend.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="D9E2F3")

spend.column_dimensions['A'].width = 18
for col in range(2, 15):
    spend.column_dimensions[get_column_letter(col)].width = 12

# ============= SHEET 3: REVENUE PROJECTION =============
rev = wb.create_sheet("Revenue Projection")

# Headers
rev.cell(row=1, column=1, value="Client").font = Font(bold=True)
for col, month in enumerate(months, 2):
    cell = rev.cell(row=1, column=col, value=month)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="538135")
    cell.alignment = Alignment(horizontal="center")

cell = rev.cell(row=1, column=14, value="Total")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="375623")

# Revenue formulas: Base + (Spend * %)
for row_idx, client in enumerate(client_names, 2):
    rev.cell(row=row_idx, column=1, value=client).font = Font(bold=True)

    for col_idx in range(2, 14):
        # Revenue = Base Retainer + (Ad Spend * % of Ad Spend)
        formula = f"='2026 Contracts'!$B${row_idx}+'2026 Ad Spend'!{get_column_letter(col_idx)}{row_idx}*'2026 Contracts'!$C${row_idx}"
        rev.cell(row=row_idx, column=col_idx, value=formula)
        rev.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0'
        rev.cell(row=row_idx, column=col_idx).font = Font(color="008000")

    # Total
    rev.cell(row=row_idx, column=14, value=f"=SUM(B{row_idx}:M{row_idx})")
    rev.cell(row=row_idx, column=14).number_format = '"$"#,##0'
    rev.cell(row=row_idx, column=14).font = Font(bold=True)

# Total row
total_row = len(client_names) + 2
rev.cell(row=total_row, column=1, value="TOTAL REVENUE").font = Font(bold=True)
rev.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="C6E0B4")
for col in range(2, 15):
    rev.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
    rev.cell(row=total_row, column=col).number_format = '"$"#,##0'
    rev.cell(row=total_row, column=col).font = Font(bold=True)
    rev.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="C6E0B4")

rev.column_dimensions['A'].width = 18
for col in range(2, 15):
    rev.column_dimensions[get_column_letter(col)].width = 12

# ============= SHEET 4: DASHBOARD =============
dash = wb.create_sheet("Dashboard")

# Title
dash.merge_cells('A1:F1')
dash.cell(row=1, column=1, value="MVR DIGITAL 2026 REVENUE ESTIMATE")
dash.cell(row=1, column=1).font = Font(bold=True, size=16)

# Key Metrics
dash.cell(row=3, column=1, value="KEY METRICS").font = Font(bold=True, size=12, color="FFFFFF")
dash.cell(row=3, column=1).fill = PatternFill("solid", fgColor="2E75B6")

metrics = [
    ("Projected Annual Revenue", "='Revenue Projection'!N7", '"$"#,##0'),
    ("Monthly Average", "='Revenue Projection'!N7/12", '"$"#,##0'),
    ("Total Ad Spend Managed", "='2026 Ad Spend'!N7", '"$"#,##0'),
    ("Effective Mgmt Fee Rate", "='Revenue Projection'!N7/'2026 Ad Spend'!N7", '0.00%'),
]

for row_idx, (label, formula, fmt) in enumerate(metrics, 4):
    dash.cell(row=row_idx, column=1, value=label)
    dash.cell(row=row_idx, column=2, value=formula)
    dash.cell(row=row_idx, column=2).number_format = fmt
    dash.cell(row=row_idx, column=2).font = Font(bold=True, size=14)

# Target Analysis
dash.cell(row=9, column=1, value="VS $150K/MO TARGET").font = Font(bold=True, size=12, color="FFFFFF")
dash.cell(row=9, column=1).fill = PatternFill("solid", fgColor="538135")

dash.cell(row=10, column=1, value="Monthly Target")
dash.cell(row=10, column=2, value=150000)
dash.cell(row=10, column=2).number_format = '"$"#,##0'
dash.cell(row=10, column=2).font = Font(color="0000FF")

dash.cell(row=11, column=1, value="Projected Avg")
dash.cell(row=11, column=2, value="=B5")
dash.cell(row=11, column=2).number_format = '"$"#,##0'

dash.cell(row=12, column=1, value="Gap")
dash.cell(row=12, column=2, value="=B11-B10")
dash.cell(row=12, column=2).number_format = '"$"#,##0;("$"#,##0)'

dash.cell(row=13, column=1, value="% of Target")
dash.cell(row=13, column=2, value="=B11/B10")
dash.cell(row=13, column=2).number_format = '0.0%'

# Monthly Breakdown
dash.cell(row=15, column=1, value="MONTHLY REVENUE").font = Font(bold=True, size=12, color="FFFFFF")
dash.cell(row=15, column=1).fill = PatternFill("solid", fgColor="C65911")

for col, month in enumerate(months, 1):
    dash.cell(row=16, column=col, value=month)
    dash.cell(row=16, column=col).font = Font(bold=True)
    dash.cell(row=16, column=col).alignment = Alignment(horizontal="center")

    # Revenue from projection
    dash.cell(row=17, column=col, value=f"='Revenue Projection'!{get_column_letter(col+1)}7")
    dash.cell(row=17, column=col).number_format = '"$"#,##0'
    dash.cell(row=17, column=col).font = Font(color="008000")

# Client breakdown
dash.cell(row=19, column=1, value="BY CLIENT (Annual)").font = Font(bold=True, size=12, color="FFFFFF")
dash.cell(row=19, column=1).fill = PatternFill("solid", fgColor="7030A0")

for row_idx, client in enumerate(client_names, 20):
    client_rev_row = row_idx - 18  # Maps to row 2-6 in Revenue Projection
    dash.cell(row=row_idx, column=1, value=client)
    dash.cell(row=row_idx, column=2, value=f"='Revenue Projection'!N{client_rev_row}")
    dash.cell(row=row_idx, column=2).number_format = '"$"#,##0'
    dash.cell(row=row_idx, column=2).font = Font(color="008000")

dash.column_dimensions['A'].width = 24
dash.column_dimensions['B'].width = 14
for col in range(3, 13):
    dash.column_dimensions[get_column_letter(col)].width = 10

# Save
output_path = "/sessions/laughing-clever-meitner/mnt/Claude/MVR_2026_Revenue_Estimate.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
