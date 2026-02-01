#!/usr/bin/env python3
"""
MVR Digital - Revenue Estimate Dashboard Builder
Creates Excel workbook for projecting monthly revenue based on % of ad spend contracts
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Create workbook
wb = Workbook()

# ============= SHEET 1: CLIENT CONFIG =============
config = wb.active
config.title = "Client Config"

# Headers
headers = ["Client", "Pricing Model", "Base Retainer", "% of Ad Spend", "Commission/Discount", "Net %", "Notes"]
for col, h in enumerate(headers, 1):
    cell = config.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center")

# Client data from 2025 invoices - Blue text for inputs
clients = [
    ["Peddle", "% of Ad Spend + Creative", 0, 0.055, 0, 0.055, "Plus creative hours at $120/hr"],
    ["SUNFLOW", "% of Attributed Revenue", 0, 0.07, 0, 0.07, "Excludes branded search; +$450 Blotout"],
    ["Kaspar & Lugay LLP", "Base + % of Ad Spend", 1000, 0.05, 0, 0.05, "Plus CC processing fees"],
    ["Le Prunier", "Base + % Attributed Rev", 3500, 0.07, -0.15, 0.0595, "15% Breef commission reduces net"],
    ["Wrensilva", "Base + % Attributed Rev", 4000, 0.05, 0, 0.05, ""],
    ["Sonsie Skin", "Base + Revenue Share", 4000, 0.08, 0, 0.08, "Variable rev share; +$350 Blotout"],
    ["Amrita", "Base + % Attributed Rev", 3500, 0.10, 0, 0.10, "Excludes branded search"],
    ["Pivot Door Company", "% of Ad Spend", 0, 0.06, 0, 0.06, ""],
    ["KIKI World", "Base + % of Ad Spend", 2000, 0.10, 0, 0.10, ""],
    ["LP Retail - Greenwich", "Flat Fee", 2000, 0, 0, 0, ""],
    ["LP Retail - Stamford", "Flat Fee", 2000, 0, 0, 0, ""],
    ["Bronx and Banco", "Flat Fee", 1000, 0, 0, 0, ""],
    ["FMW Fasteners", "Flat Fee", 1000, 0, 0, 0, "Plus CC processing fee"],
    ["[New Client 1]", "Base + % of Ad Spend", 0, 0.05, 0, 0.05, "Update with actual"],
    ["[New Client 2]", "Base + % of Ad Spend", 0, 0.05, 0, 0.05, "Update with actual"],
]

for row_idx, client in enumerate(clients, 2):
    for col_idx, val in enumerate(client, 1):
        cell = config.cell(row=row_idx, column=col_idx, value=val)
        if col_idx in [3, 4, 5]:  # Input columns
            cell.font = Font(color="0000FF")  # Blue for inputs
        if col_idx == 6:  # Net % is formula
            cell.font = Font(color="000000")
            if row_idx >= 2:
                # Net % = Ad Spend % * (1 + Commission/Discount)
                config.cell(row=row_idx, column=6, value=f"=D{row_idx}*(1+E{row_idx})")

# Format columns
config.column_dimensions['A'].width = 22
config.column_dimensions['B'].width = 22
config.column_dimensions['C'].width = 14
config.column_dimensions['D'].width = 14
config.column_dimensions['E'].width = 18
config.column_dimensions['F'].width = 10
config.column_dimensions['G'].width = 35

# Format percentages
for row in range(2, len(clients) + 2):
    config.cell(row=row, column=3).number_format = '"$"#,##0'
    config.cell(row=row, column=4).number_format = '0.0%'
    config.cell(row=row, column=5).number_format = '0.0%'
    config.cell(row=row, column=6).number_format = '0.0%'

# ============= SHEET 2: AD SPEND INPUT =============
spend = wb.create_sheet("Ad Spend Input")

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
client_names = [c[0] for c in clients]

# Header row
spend.cell(row=1, column=1, value="Client").font = Font(bold=True)
for col, month in enumerate(months, 2):
    cell = spend.cell(row=1, column=col, value=f"{month} 2026")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center")

# Total column
cell = spend.cell(row=1, column=14, value="Total")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="1F4E79")
cell.alignment = Alignment(horizontal="center")

# Client rows with sample data
sample_spend = {
    "Peddle": [700000, 720000, 750000, 780000, 800000, 850000, 900000, 950000, 900000, 850000, 900000, 950000],
    "SUNFLOW": [20000, 25000, 40000, 60000, 80000, 100000, 120000, 100000, 60000, 30000, 40000, 80000],
    "Kaspar & Lugay LLP": [80000, 85000, 90000, 95000, 100000, 100000, 100000, 100000, 100000, 95000, 100000, 105000],
    "Wrensilva": [40000, 45000, 50000, 55000, 60000, 70000, 80000, 90000, 100000, 120000, 130000, 150000],
    "Sonsie Skin": [30000, 35000, 40000, 45000, 50000, 60000, 80000, 70000, 60000, 70000, 80000, 100000],
    "Amrita": [20000, 22000, 25000, 28000, 30000, 35000, 40000, 45000, 40000, 35000, 40000, 50000],
    "Pivot Door Company": [25000, 27000, 30000, 32000, 35000, 38000, 40000, 38000, 35000, 33000, 35000, 38000],
    "KIKI World": [15000, 18000, 20000, 22000, 25000, 28000, 30000, 28000, 25000, 22000, 20000, 25000],
    "Le Prunier": [30000, 35000, 40000, 45000, 50000, 55000, 60000, 55000, 50000, 45000, 50000, 60000],
}

for row_idx, client in enumerate(client_names, 2):
    cell = spend.cell(row=row_idx, column=1, value=client)
    cell.font = Font(bold=True)

    # Add sample spend data or zeros
    for col_idx, month in enumerate(months, 2):
        val = sample_spend.get(client, [0]*12)[col_idx-2] if client in sample_spend else 0
        cell = spend.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(color="0000FF")  # Blue for inputs
        cell.number_format = '"$"#,##0'

    # Total formula
    spend.cell(row=row_idx, column=14, value=f"=SUM(B{row_idx}:M{row_idx})")
    spend.cell(row=row_idx, column=14).number_format = '"$"#,##0'
    spend.cell(row=row_idx, column=14).font = Font(bold=True)

# Monthly totals row
total_row = len(client_names) + 2
spend.cell(row=total_row, column=1, value="TOTAL AD SPEND").font = Font(bold=True)
for col in range(2, 15):
    spend.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
    spend.cell(row=total_row, column=col).number_format = '"$"#,##0'
    spend.cell(row=total_row, column=col).font = Font(bold=True)
    spend.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E2EFDA")

# Format columns
spend.column_dimensions['A'].width = 22
for col in range(2, 15):
    spend.column_dimensions[get_column_letter(col)].width = 12

# ============= SHEET 3: REVENUE PROJECTION =============
proj = wb.create_sheet("Revenue Projection")

# Header
proj.cell(row=1, column=1, value="Client").font = Font(bold=True)
for col, month in enumerate(months, 2):
    cell = proj.cell(row=1, column=col, value=f"{month} 2026")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="538135")
    cell.alignment = Alignment(horizontal="center")

cell = proj.cell(row=1, column=14, value="Total")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill("solid", fgColor="375623")
cell.alignment = Alignment(horizontal="center")

# Revenue formulas for each client
for row_idx, client in enumerate(client_names, 2):
    cell = proj.cell(row=row_idx, column=1, value=client)
    cell.font = Font(bold=True)

    for col_idx in range(2, 14):
        # Revenue = Base Retainer + (Ad Spend * Net %)
        # Base is in Client Config column C, Net % in column F
        # Ad Spend is in Ad Spend Input same row/column
        formula = f"='Client Config'!$C${row_idx}+'Ad Spend Input'!{get_column_letter(col_idx)}{row_idx}*'Client Config'!$F${row_idx}"
        proj.cell(row=row_idx, column=col_idx, value=formula)
        proj.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0'
        proj.cell(row=row_idx, column=col_idx).font = Font(color="008000")  # Green for cross-sheet refs

    # Total formula
    proj.cell(row=row_idx, column=14, value=f"=SUM(B{row_idx}:M{row_idx})")
    proj.cell(row=row_idx, column=14).number_format = '"$"#,##0'
    proj.cell(row=row_idx, column=14).font = Font(bold=True)

# Monthly totals
total_row = len(client_names) + 2
proj.cell(row=total_row, column=1, value="TOTAL REVENUE").font = Font(bold=True)
for col in range(2, 15):
    proj.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})")
    proj.cell(row=total_row, column=col).number_format = '"$"#,##0'
    proj.cell(row=total_row, column=col).font = Font(bold=True)
    proj.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="C6E0B4")

# Format columns
proj.column_dimensions['A'].width = 22
for col in range(2, 15):
    proj.column_dimensions[get_column_letter(col)].width = 12

# ============= SHEET 4: DASHBOARD =============
dash = wb.create_sheet("Dashboard")

# Title
dash.merge_cells('A1:G1')
dash.cell(row=1, column=1, value="MVR DIGITAL - 2026 REVENUE ESTIMATE DASHBOARD")
dash.cell(row=1, column=1).font = Font(bold=True, size=16)
dash.cell(row=1, column=1).alignment = Alignment(horizontal="center")

# Summary Section
dash.cell(row=3, column=1, value="ANNUAL SUMMARY").font = Font(bold=True, size=12)
dash.cell(row=3, column=1).fill = PatternFill("solid", fgColor="2E75B6")
dash.cell(row=3, column=1).font = Font(bold=True, color="FFFFFF", size=12)

metrics = [
    ("Projected Annual Revenue", "='Revenue Projection'!N17", '"$"#,##0'),
    ("Monthly Average", "='Revenue Projection'!N17/12", '"$"#,##0'),
    ("Total Ad Spend Managed", "='Ad Spend Input'!N17", '"$"#,##0'),
    ("Avg Effective Rate", "='Revenue Projection'!N17/'Ad Spend Input'!N17", '0.0%'),
]

for row_idx, (label, formula, fmt) in enumerate(metrics, 4):
    dash.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
    dash.cell(row=row_idx, column=2, value=formula)
    dash.cell(row=row_idx, column=2).number_format = fmt
    dash.cell(row=row_idx, column=2).font = Font(bold=True, size=14)

# Target comparison
dash.cell(row=9, column=1, value="VS TARGETS").font = Font(bold=True, size=12)
dash.cell(row=9, column=1).fill = PatternFill("solid", fgColor="538135")
dash.cell(row=9, column=1).font = Font(bold=True, color="FFFFFF", size=12)

dash.cell(row=10, column=1, value="Monthly Target")
dash.cell(row=10, column=2, value=150000)
dash.cell(row=10, column=2).number_format = '"$"#,##0'
dash.cell(row=10, column=2).font = Font(color="0000FF")

dash.cell(row=11, column=1, value="Projected Monthly Avg")
dash.cell(row=11, column=2, value="=B5")
dash.cell(row=11, column=2).number_format = '"$"#,##0'

dash.cell(row=12, column=1, value="Gap to Target")
dash.cell(row=12, column=2, value="=B10-B11")
dash.cell(row=12, column=2).number_format = '"$"#,##0;("$"#,##0)'

dash.cell(row=13, column=1, value="% of Target")
dash.cell(row=13, column=2, value="=B11/B10")
dash.cell(row=13, column=2).number_format = '0.0%'

# Monthly breakdown
dash.cell(row=15, column=1, value="MONTHLY REVENUE PROJECTION").font = Font(bold=True, size=12)
dash.cell(row=15, column=1).fill = PatternFill("solid", fgColor="C65911")
dash.cell(row=15, column=1).font = Font(bold=True, color="FFFFFF", size=12)

for col, month in enumerate(months, 1):
    dash.cell(row=16, column=col, value=f"{month}")
    dash.cell(row=16, column=col).font = Font(bold=True)
    dash.cell(row=16, column=col).alignment = Alignment(horizontal="center")

    # Reference to Revenue Projection totals row
    dash.cell(row=17, column=col, value=f"='Revenue Projection'!{get_column_letter(col+1)}17")
    dash.cell(row=17, column=col).number_format = '"$"#,##0'
    dash.cell(row=17, column=col).font = Font(color="008000")

# Target line
dash.cell(row=18, column=1, value="Target").font = Font(bold=True)
for col in range(1, 13):
    dash.cell(row=18, column=col, value="=$B$10")
    dash.cell(row=18, column=col).number_format = '"$"#,##0'
    dash.cell(row=18, column=col).font = Font(color="FF0000")

# Client contribution section
dash.cell(row=20, column=1, value="TOP CLIENTS (Projected Annual)").font = Font(bold=True, size=12)
dash.cell(row=20, column=1).fill = PatternFill("solid", fgColor="7030A0")
dash.cell(row=20, column=1).font = Font(bold=True, color="FFFFFF", size=12)

top_clients = ["Peddle", "SUNFLOW", "Kaspar & Lugay LLP", "Wrensilva", "Sonsie Skin"]
for row_idx, client in enumerate(top_clients, 21):
    dash.cell(row=row_idx, column=1, value=client)
    # Reference their total from Revenue Projection
    client_row = client_names.index(client) + 2 if client in client_names else 2
    dash.cell(row=row_idx, column=2, value=f"='Revenue Projection'!N{client_row}")
    dash.cell(row=row_idx, column=2).number_format = '"$"#,##0'
    dash.cell(row=row_idx, column=2).font = Font(color="008000")

# Format dashboard columns
dash.column_dimensions['A'].width = 28
dash.column_dimensions['B'].width = 15
for col in range(3, 13):
    dash.column_dimensions[get_column_letter(col)].width = 10

# Save
output_path = "/sessions/laughing-clever-meitner/mnt/Claude/MVR_Revenue_Estimate_2026.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
